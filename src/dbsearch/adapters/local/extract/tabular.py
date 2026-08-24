"""CSV (stdlib) / XLSX (openpyxl) -> one segment per data row. Header prepended as
key: value pairs for retrievability. Row number n counts from the header as row 1."""
from __future__ import annotations

import csv
import io

from dbsearch.core.models import Segment


def _row_text(header: list[str], row: list) -> str:
    if header:
        return "; ".join(f"{h}: {v}" for h, v in zip(header, row) if str(v).strip())
    return "; ".join(str(v) for v in row if str(v).strip())


def segment_csv(data: bytes) -> list[Segment]:
    rows = list(csv.reader(io.StringIO(data.decode("utf-8", "ignore"))))
    if not rows:
        return []
    header = rows[0]
    out: list[Segment] = []
    for n, row in enumerate(rows[1:], start=2):  # header = row 1
        line = _row_text(header, row)
        if line.strip():
            out.append(Segment(text=line, locator={"kind": "row", "n": n}))
    return out


def segment_xlsx(data: bytes) -> list[Segment]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    multi = len(wb.worksheets) > 1
    out: list[Segment] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c) if c is not None else "" for c in rows[0]]
        for n, row in enumerate(rows[1:], start=2):
            vals = [c if c is not None else "" for c in row]
            line = _row_text(header, vals)
            if line.strip():
                loc = {"kind": "row", "n": n}
                if multi:
                    loc["sheet"] = ws.title
                out.append(Segment(text=line, locator=loc))
    return out
